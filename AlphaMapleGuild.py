from datetime import datetime
import cv2
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import time
from PIL import ImageGrab
import os
import hashlib

# =============================================================================
# 신규 길드 UI 대응 골격 (2026 리디자인)
# -----------------------------------------------------------------------------
# 변경 요약:
#   1) 부드러운 스크롤    -> 고정 그리드 폐기. 연속 캡처 + 동적 행 탐지 + 해시 dedup
#   2) 미세 반투명        -> 측정 결과 패널 사실상 불투명. exact-color 매칭 유지
#   3) 온라인 괄호 텍스트 -> 괄호 파싱 안 함. '보이는 형태' 전체를 해시하고
#                            엑셀 [닉네임 정보] 시트의 '대표캐릭' 칼럼으로 수동 매핑
#   4) 점수는 계정 단위 1개 -> 합산 없음 (계정당 1행)
#
# 픽셀 의존 값은 전부 아래 상수 블록으로 분리. 'TODO(픽셀)' 만 채우면 동작.
# 숫자 폰트(numbers.png / number_1000.png)는 변경 없음(구현 직전 픽셀 대조로 확인).
# =============================================================================

# --- 템플릿 에셋 -------------------------------------------------------------
GUILD_TEMPLATE = cv2.imread(r"data\guild.png", cv2.IMREAD_GRAYSCALE)
GUILD_MASK = cv2.imread(r"data\mask.png", cv2.IMREAD_GRAYSCALE)
NUMBERS = cv2.imread(r"data\numbers.png", cv2.IMREAD_GRAYSCALE)
THOUSAND = cv2.imread(r"data\number_1000.png", cv2.IMREAD_GRAYSCALE) * 255  # uint8 오버플로로 0/1 화

GUILD_H, GUILD_W = GUILD_TEMPLATE.shape          # 신규 창 ≈ (683, 652)
NUMBERS = NUMBERS.reshape(8, 10, 5).swapaxes(0, 1) * 255  # (10, 8, 5), 0/1
DIGIT_H, DIGIT_W = 8, 5                            # 숫자 글리프 크기 (폰트 동일)

# --- 텍스트 색 (ImageGrab = RGB) ---------------------------------------------
# 데이터 영역 텍스트는 두 색뿐 (측정 확인):
#   온라인  = (255,255,255)  닉네임+숫자 모두 흰색
#   오프라인 = (171,181,187) 닉네임+숫자 모두 회색
# 이진화 시 둘 다 1로 normalize. (사이드바 dim색 108,119,133 은 데이터영역 밖이라 제외)
TEXT_COLORS = [(255, 255, 255), (171, 181, 187)]

# --- 레이아웃 ----------------------------------------------------------------
# 좌표계:  y(세로) = 창 좌상단 기준  /  x(가로) = 데이터영역 좌단(DATA_X0) 기준.
#          x 를 슬라이스할 때는 항상 DATA_X0 를 더한다.
DATA_X0 = 121                # 데이터 영역(시안 하이라이트) 좌단 = 모든 컬럼 x 의 원점

# 행 격자 (창 기준 y)
ROW_PITCH   = 26             # 행 간격(px)
LIST_TOP    = 137            # 목록 상단 y (= 첫 행 top). 위로 잘린 행 버림
LIST_BOTTOM = 626            # 목록 하단 y. 아래로 잘린 행 버림
ROW_TEXT_H  = 16             # 행 top ~ 텍스트 최하단까지 높이 (잘림 판정용).
                             #   잘림판정: (y + ROW_TEXT_H - 1) <= LIST_BOTTOM
                             #   (하이라이트 전체높이 21 이 아님 — 빈 꼬리 잘려도 텍스트 보이면 채택)

# 행 내부 세로 오프셋 (행 top 기준 y)
DIGIT_DY = 6                 # 숫자 글리프 시작 y — 모든 점수 컬럼 공통 (6:14, 8px)
NICK_DY0, NICK_DY1 = 5, 16   # 닉네임 글리프 y 범위 (해시 대상)

# 컬럼 가로 위치 (DATA_X0 기준 x)
NICK_X0, NICK_X1 = 30, 100                  # 닉네임 (왼쪽+괄호까지 '보이는 형태' 전체)
MISSION_TWO_DIGITS = [(330, 10), (336, 1)]  # 주미 두 자리 "10"
MISSION_ONE_DIGIT  = [(333, 1)]             # 주미 한 자리 (가운데 정렬)
FLAG_THOUSAND_X    = 460                    # 플래그 "1,000" 템플릿 좌측 x (0/1000 판정)
CANAL_PROBE_X = [379, 382, 385, 389, 392, 395]   # 수로 최상위 자리 시작 x 후보 (긴 수 우선)
CANAL_LAYOUTS = {                                # 자리수별 [(x, 배수), ...]
    379: [(379, 100000), (385, 10000), (391, 1000), (400, 100), (406, 10), (412, 1)],
    382: [(382, 10000), (388, 1000), (397, 100), (403, 10), (409, 1)],
    385: [(385, 1000), (394, 100), (400, 10), (406, 1)],
    389: [(389, 100), (395, 10), (401, 1)],
    392: [(392, 10), (398, 1)],
    395: [(395, 1)],
}

# 탭 확인은 제거: '컨텐츠 참여 현황' 탭(청록 강조)이 guild.png/mask.png 매칭 영역에
# 포함되므로, 다른 탭이면 매칭 오차가 올라가 창 탐색 단계에서 자동으로 걸러진다.

# --- 스캔 동작 ---------------------------------------------------------------
POLL_INTERVAL   = 0.50   # 연속 캡처 간격(초). 부드러운 스크롤을 프레임으로 훑음
IDLE_TIMEOUT    = 10     # 신규 행이 이 시간(초) 동안 없으면 스캔 종료
BLANK_MIN_PIXELS = 3     # 닉네임 영역 텍스트 픽셀이 이 이하이면 빈 행으로 간주
CONFIRM_FRAMES  = 2      # 같은 해시가 N프레임 연속 관측돼야 확정(커서 가림/모션 방어)


# =============================================================================
# 캡처 / 이진화
# =============================================================================
def grab_mask(gx, gy):
    """창 영역을 캡처해 텍스트색만 1로 만든 이진 마스크(H, W) 반환."""
    cap = np.array(ImageGrab.grab(bbox=(gx, gy, gx + GUILD_W, gy + GUILD_H), all_screens=True))
    mask = np.zeros(cap.shape[:2], dtype=bool)
    for color in TEXT_COLORS:
        mask |= np.all(cap == color, axis=-1)
    return mask.astype('u1')


# =============================================================================
# 동적 행 탐지
#   고정 그리드를 버리고, 닉네임 컬럼의 텍스트 밴드로 행 위상(phase)을 추정.
#   pitch는 일정하므로 phase만 잡으면 모든 행을 anchor + k*pitch 로 인덱싱.
# =============================================================================
def row_anchors(mask):
    """완전히 보이는 행들의 '앵커 y'(행 top = 시안 하이라이트 top 기준) 리스트 반환.
    모든 오프셋(NICK_*, DIGIT_DY, 점수)은 이 행 top 기준으로 측정한다."""
    col = mask[:, DATA_X0 + NICK_X0:DATA_X0 + NICK_X1]
    present = np.where(col.sum(axis=1) > BLANK_MIN_PIXELS)[0]
    if len(present) == 0:
        return []

    # 텍스트 밴드 시작점들로 행 위상 추정 (글리프 잡음에 강하도록 중앙값 사용).
    # 닉네임 텍스트는 행 top 보다 NICK_DY0 만큼 아래에서 시작(측정: 행top%26=7, 닉텍스트%26=12).
    # 따라서 NICK_DY0 를 빼서 앵커를 '행 top'에 맞춘다.
    starts = present[np.concatenate(([True], np.diff(present) > 1))]
    phase = (int(np.median(starts % ROW_PITCH)) - NICK_DY0) % ROW_PITCH

    anchors = []
    y = phase
    while y + ROW_TEXT_H <= GUILD_H:
        # 목록 영역 안에 '완전히' 들어오는 행만 채택 (위/아래 잘린 행 제외).
        # 하이라이트는 [y, y+ROW_TEXT_H-1] inclusive 이므로 bottom 픽셀에 -1.
        if LIST_TOP <= y and (y + ROW_TEXT_H - 1) <= LIST_BOTTOM:
            anchors.append(y)
        y += ROW_PITCH
    return anchors


# =============================================================================
# 숫자 판독 (모두 행 앵커 기준 상대좌표)
# =============================================================================
def read_digits(mask, anchor, digit_positions):
    """digit_positions: [(x, multiplier), ...] -> 정수. 매칭 실패 자리는 0."""
    total = 0
    y0 = anchor + DIGIT_DY
    for x, mult in digit_positions:
        cell = mask[y0:y0 + DIGIT_H, DATA_X0 + x:DATA_X0 + x + DIGIT_W]
        idx = np.where(np.all(cell[np.newaxis] == NUMBERS, axis=(1, 2)))[0]
        if len(idx):
            total += int(idx[0]) * mult
    return total


def read_mission(mask, anchor):
    # 두 자리("10")와 한 자리(가운데 정렬) 중 실제로 잡히는 쪽만 값이 더해짐
    return read_digits(mask, anchor, MISSION_TWO_DIGITS) + read_digits(mask, anchor, MISSION_ONE_DIGIT)


def read_flag(mask, anchor):
    # 0 또는 1,000 뿐: "1,000" 템플릿이 매칭되면 1000, 아니면 0.
    # 높이는 THOUSAND 템플릿(콤마 포함 9px)에서 자동 -> 콤마 +1px 를 수동으로 맞출 필요 없음.
    y0 = anchor + DIGIT_DY
    fx = DATA_X0 + FLAG_THOUSAND_X
    cell = mask[y0:y0 + THOUSAND.shape[0], fx:fx + THOUSAND.shape[1]]
    return 1000 if cell.shape == THOUSAND.shape and np.all(cell == THOUSAND) else 0


def read_canal(mask, anchor):
    # 첫 숫자가 잡히는 후보 x로 자리수를 판별한 뒤 해당 레이아웃으로 판독
    y0 = anchor + DIGIT_DY
    for start_x in CANAL_PROBE_X:           # 왼쪽(큰 자리수) 우선
        cell = mask[y0:y0 + DIGIT_H, DATA_X0 + start_x:DATA_X0 + start_x + DIGIT_W]
        if np.any(np.all(cell[np.newaxis] == NUMBERS, axis=(1, 2))):
            return read_digits(mask, anchor, CANAL_LAYOUTS[start_x])
    return 0


def nickname_crop(mask, anchor):
    return mask[anchor + NICK_DY0:anchor + NICK_DY1, DATA_X0 + NICK_X0:DATA_X0 + NICK_X1]


# =============================================================================
# 1) 창 위치 탐색
# =============================================================================
print("전체 화면을 캡처해 길드 창의 위치를 찾습니다. 모니터의 해상도가 높을수록 오래 걸릴 수 있습니다.")

capture = cv2.cvtColor(np.array(ImageGrab.grab(all_screens=True)), cv2.COLOR_BGR2GRAY)
best_error, _, best_location, _ = cv2.minMaxLoc(
    cv2.matchTemplate(capture, GUILD_TEMPLATE, cv2.TM_SQDIFF_NORMED, mask=GUILD_MASK))

print(f"이미지 오차 점수: {best_error:.3f}, 좌상단 픽셀 좌표 (x, y): {best_location}")
if best_error > 0.1:
    input("오차는 대체로 0.01 이하입니다. [Enter] 키를 입력해 무시하거나 다시 시작해 주세요.")

guild_x, guild_y = best_location

# =============================================================================
# 2) 연속 캡처 스캔 루프
#   - 매 프레임 동적 행 탐지 -> 완전히 보이는 행만 판독
#   - 해시 dedup 으로 중복/재방문 제거
#   - 모션/커서 가림 프레임은 exact-match 실패로 자연 필터 + CONFIRM_FRAMES 로 추가 방어
# =============================================================================
if not os.path.exists('nick_hash'):
    os.makedirs('nick_hash')

written_records = {}     # hash -> (mission, canal, flag)  확정된 행
pending = {}             # hash -> 연속 관측 횟수 (CONFIRM_FRAMES 도달 시 확정)
scanned_query = 0
last_progress = time.time()

print("스크롤을 시작해 주세요. 닉네임/숫자를 커서로 가리지 않도록 주의해 주세요.")
while True:
    mask = grab_mask(guild_x, guild_y)
    seen_this_frame = set()

    for anchor in row_anchors(mask):
        nick = nickname_crop(mask, anchor)
        if int(nick.sum()) <= BLANK_MIN_PIXELS:
            continue  # 빈 행

        h = hashlib.sha256(nick.tobytes()).hexdigest()
        seen_this_frame.add(h)
        if h in written_records:
            continue

        # 모션/가림 방어: 같은 해시가 CONFIRM_FRAMES 만큼 연속 관측돼야 확정
        pending[h] = pending.get(h, 0) + 1
        if pending[h] < CONFIRM_FRAMES:
            continue

        record = (read_mission(mask, anchor), read_canal(mask, anchor), read_flag(mask, anchor))
        written_records[h] = record
        cv2.imwrite(f'nick_hash/{h}.png', 255 - nick * 255)

    # 이번 프레임에 안 보인 미확정 항목은 카운트 리셋(스크롤로 지나가버린 행)
    for h in list(pending):
        if h not in seen_this_frame or h in written_records:
            pending.pop(h, None)

    newly = len(written_records) - scanned_query
    if newly > 0:
        scanned_query = len(written_records)
        print(f"\r{newly}명 추가, 누적 {scanned_query}명 기록")
        last_progress = time.time()
    elif time.time() - last_progress > IDLE_TIMEOUT:
        print("\r시간이 초과되어 스캔을 종료합니다.", end='\n\n')
        break
    else:
        print(f"\r스크롤을 넘겨주세요. {IDLE_TIMEOUT - (time.time() - last_progress):.1f}초간 새 항목이 없으면 종료합니다.", end='')

    time.sleep(POLL_INTERVAL)
# 스캔 종료


# =============================================================================
# 3) 엑셀 기록
#   [닉네임 정보] : A=이미지  B=해시(숨김)  C=닉네임(수동)  D=대표캐릭(수동, 신규)
#   일일 시트     : 신원 = 대표캐릭 우선, 없으면 닉네임, 둘 다 없으면 "???"
# =============================================================================
try:  # 기존 워크북 로드
    wb = load_workbook('result.xlsx')
    w0 = wb.worksheets[0]
    # B=해시, C=닉네임, D=대표캐릭
    known = {row[0]: (row[1], row[2])
             for row in w0.iter_rows(min_row=2, max_row=w0.max_row, min_col=2, max_col=4, values_only=True)}
except FileNotFoundError:  # 신규 생성
    wb = Workbook()
    w0 = wb.worksheets[0]
    w0.title = "닉네임 정보"
    w0.column_dimensions['B'].hidden = True
    w0.cell(row=1, column=1).value = '닉네임 이미지'
    w0.cell(row=1, column=2).value = '닉네임 해시'
    w0.cell(row=1, column=3).value = '닉네임(수동입력)'
    w0.cell(row=1, column=4).value = '대표캐릭(수동입력)'
    known = {}

# 오늘 시트
ws = wb.create_sheet(title=datetime.now().strftime("%y%m%d_%H%M"))
ws.cell(row=1, column=1).value = '닉네임 해시'
ws.column_dimensions['A'].hidden = True
ws.cell(row=1, column=2).value = '대표캐릭(자동으로 채워짐)'
ws.cell(row=1, column=3).value = '주간미션'
ws.cell(row=1, column=4).value = '지하 수로'
ws.cell(row=1, column=5).value = '플래그 레이스'

naive_query = 0
for current_row, (hash_, (m, s, f)) in enumerate(written_records.items(), start=2):
    ws.cell(row=current_row, column=1).value = hash_
    ws.cell(row=current_row, column=3).value = m
    ws.cell(row=current_row, column=4).value = s
    ws.cell(row=current_row, column=5).value = f

    if hash_ in known:
        nick, rep = known.pop(hash_)
        # 대표캐릭 우선, 없으면 닉네임. 둘 다 수기로 채워져 있으면 정적 기록.
        resolved = rep or nick
        if resolved is not None:
            ws.cell(row=current_row, column=2).value = resolved
            continue
    else:
        naive_query += 1
        max_r = w0.max_row + 1
        w0.add_image(Image(f'nick_hash/{hash_}.png'), anchor=f'A{max_r}')
        w0.cell(row=max_r, column=2).value = hash_

    # 대표캐릭(D) 우선, 비면 닉네임(C), 둘 다 비면 "???"
    rep_f = f"VLOOKUP(A{current_row},'닉네임 정보'!B:D,3,FALSE)"
    nick_f = f"VLOOKUP(A{current_row},'닉네임 정보'!B:D,2,FALSE)"
    ws.cell(row=current_row, column=2).value = (
        f'=IFERROR(IF({rep_f}<>"",{rep_f},IF({nick_f}<>"",{nick_f},"???")),"???")')

# 누락/신규 경고
warn = False
if len(known) > 0:
    print("============================경고============================")
    print(f"기존에 존재한 닉네임 중 총 {len(known)}개가 스캔되지 않았습니다.")
    print("스캔 도중 마우스가 닉네임을 가렸거나, 닉네임을 변경했거나, 길드를 탈퇴하였을 수 있습니다.")
    print("누락된 닉네임 중, 수기로 기록되어 있던 것은 다음과 같습니다.")
    print([nick for nick, rep in known.values() if nick or rep])
    warn = True
if naive_query > 0:
    if not warn:
        print("============================경고============================")
    print(f"기존에 없던 {naive_query}개의 닉네임이 추가되었습니다.")
    print("프로그램 종료 후 [닉네임 정보] 시트에서 닉네임/대표캐릭을 수동으로 입력해 주세요.")
    print("같은 계정의 접속캐릭은 [대표캐릭] 칼럼에 대표명을 적어 한 사람으로 묶을 수 있습니다.")
    warn = True

if warn:
    input("변동사항을 적용하려면 [Enter]를 입력하세요.\n============================================================\n")
else:
    print(f"총 {scanned_query}개의 닉네임이 스캔되었으며, 기존 닉네임 중 변동된 닉네임은 없습니다.")

while True:
    try:
        wb.save('result.xlsx')
        input("result.xlsx 파일이 저장되었습니다. [Enter]를 입력하거나 창을 닫아 종료하세요.")
        break
    except PermissionError:
        input("권한 오류가 발생하였습니다. [Enter]를 입력해 재시도합니다. 주로 result.xlsx가 열려 있는 경우 발생합니다.")
